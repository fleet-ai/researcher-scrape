#!/usr/bin/env python3
"""Deep-verify the top-N candidates per position via web search.

Takes expand.py's per-position CSVs (the wide candidate pool) and produces
recruiter-ready shortlists in the proven format:

    #, Name, Career Stage, Key Work, Personal Email, Website, Recruitable?

For each candidate an LLM with live web search resolves what OpenAlex cannot:
current employer (with transitions, e.g. "PhD 24, MIT -> OpenAI"), PhD
year/status, personal email from their homepage, personal website, a one-line
key-work gloss with venues, and a recruitability call with a reason.

Usage:
    python verify_shortlist.py                          # all tabs, top 40 each
    python verify_shortlist.py --top 25 --tabs world_models
    python verify_shortlist.py --input-dir data/expand_output
"""

import argparse
import csv
import json
import logging
import os
import re
import time
from pathlib import Path

import requests
from pydantic import BaseModel, Field, ValidationError

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent / "data"
VERIFY_CACHE_PATH = DATA_DIR / "verify_cache.json"
HUNT_CACHE_PATH = DATA_DIR / "email_hunt_cache.json"

OPENROUTER_API = "https://openrouter.ai/api/v1/chat/completions"
# ":online" enables OpenRouter's web-search augmentation on any model
VERIFY_MODEL = os.environ.get("VERIFY_MODEL", "anthropic/claude-sonnet-4.5:online")

STAGE_RANK = {
    "graduating_phd": 0, "recent_grad": 1, "junior_industry": 2,
    "mid_industry": 3, "unknown": 4, "": 4,
    "early_phd": 5, "founder": 6, "senior": 7, "professor": 8,
}
RECRUIT_RANK = {"Yes": 0, "": 1, "Stretch": 2, "Unlikely": 3}


class Verification(BaseModel):
    """What the web-search model must return for one candidate."""
    identity_confirmed: bool = Field(description="True if web results clearly match this person (papers line up)")
    career_stage: str = Field(default="", description='e.g. "Graduating PhD (4th yr), Tsinghua" or "Recent grad (PhD 24, MIT) -> OpenAI"')
    current_employer: str = Field(default="")
    key_work: str = Field(default="", description='1-2 works with venue+year and a short gloss, e.g. "TD-MPC2 (ICLR 24); Holodeck (CVPR 24) - 3D env generation from language"')
    personal_email: str = Field(default="", description="Personal (gmail etc.) preferred over institutional; empty if not found")
    website: str = Field(default="", description="Personal site domain, no https:// prefix")
    recruitable: str = Field(default="", description='"Yes", "Maybe" or "Unlikely" with an optional reason, e.g. "Unlikely - at OpenAI" or "Yes - just switched labs"')
    notes: str = Field(default="")


VERIFY_PROMPT = """Research this ML researcher on the web (personal website, Google Scholar, LinkedIn, X/Twitter, lab pages) and report their CURRENT status for a recruiting shortlist.

Name: {name}
Last known institution (may be stale): {institution}
Known papers: {papers}
h-index: {h_index}

We are hiring for: {category}. Focus the key-work gloss on that.

Rules:
- Confirm identity by matching the papers above. If you cannot confirm, set identity_confirmed=false and leave other fields empty.
- career_stage must be specific: PhD year + school + current employer with an arrow for transitions, e.g. "Recent grad (PhD 24, UPenn) -> AI2" or "Graduating PhD (final yr), Stanford".
- key_work: at most 2 works, one line, under 140 characters total, e.g. "TD-MPC2 (ICLR 24), Multitask World Models (ICLR 26)" or "Holodeck (CVPR 24) - 3D env generation from language".
- personal_email: prefer a personal address found on their homepage/CV; never invent one.
- recruitable: "Yes" for grad students / academics / startup folks; "Maybe" for big-lab (Meta FAIR, DeepMind, NVIDIA) juniors; "Unlikely - at X" for OpenAI/Anthropic; add a short reason when it helps ("Yes - just switched labs").
- Respond with ONLY a JSON object matching this schema, no prose:
{{"identity_confirmed": bool, "career_stage": str, "current_employer": str, "key_work": str, "personal_email": str, "website": str, "recruitable": str, "notes": str}}"""


def _normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s.lower().strip())


def _load_json(path: Path) -> dict:
    if path.exists():
        try:
            return json.loads(path.read_text())
        except Exception:
            pass
    return {}


def _parse_json_obj(content: str) -> dict:
    content = content.strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
    # Some models wrap the object in prose; grab the outermost braces
    m = re.search(r"\{.*\}", content, re.DOTALL)
    return json.loads(m.group(0) if m else content)


def verify_candidate(row: dict, category: str, api_key: str) -> Verification | None:
    prompt = VERIFY_PROMPT.format(
        name=row["name"],
        institution=row.get("institution") or "unknown",
        papers=(row.get("key_papers") or "")[:400],
        h_index=row.get("h_index", ""),
        category=category,
    )
    for attempt in range(3):
        try:
            resp = requests.post(
                OPENROUTER_API,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={"model": VERIFY_MODEL,
                      "messages": [{"role": "user", "content": prompt}],
                      "temperature": 0, "max_tokens": 1500},
                timeout=180,
            )
            if resp.status_code == 429:
                log.warning("OpenRouter rate limited, sleeping 15s")
                time.sleep(15)
                continue
            if resp.status_code != 200:
                log.warning(f"OpenRouter {resp.status_code}: {resp.text[:200]}")
                return None
            content = resp.json()["choices"][0]["message"]["content"]
            return Verification(**_parse_json_obj(content))
        except (json.JSONDecodeError, ValidationError) as e:
            log.warning(f"Parse error for {row['name']} (attempt {attempt + 1}): {e}")
        except Exception as e:
            log.warning(f"Request failed for {row['name']} (attempt {attempt + 1}): {e}")
            time.sleep(5)
    return None


def rank_rows(rows: list[dict]) -> list[dict]:
    """Hirability sort: recruitable Yes first, then career-stage order, then h desc."""
    def key(r):
        return (RECRUIT_RANK.get(r.get("recruitable", ""), 1),
                STAGE_RANK.get(r.get("career_stage", ""), 4),
                -int(r.get("h_index") or 0))
    return sorted(rows, key=key)


SHORTLIST_FIELDS = ["#", "Name", "Career Stage", "Key Work", "Personal Email", "Website", "Recruitable?"]


def process_tab(csv_path: Path, category: str, top_n: int, api_key: str,
                cache: dict, hunt_cache: dict, out_dir: Path,
                max_verify: int = 0) -> list[dict]:
    with open(csv_path) as f:
        rows = list(csv.DictReader(f))
    # Verify-then-cut, not cut-then-verify: walk the ranked list and keep
    # going until top_n slots are filled with identity-confirmed Yes/Maybe
    # candidates. Unconfirmed identities and verified-Unlikely rows consume
    # verification budget but never a slot. max_verify bounds the walk so a
    # tab full of Unlikelies can't burn the whole pool (0 = 4x top_n).
    ranked = rank_rows(rows)
    cap = max_verify or 4 * top_n
    ranked = ranked[:cap]
    log.info(f"[{category}] filling {top_n} Yes/Maybe slots from {len(rows)} "
             f"candidates (verify cap {len(ranked)})")

    out_rows = []
    n_unconfirmed = n_unlikely = 0
    for i, row in enumerate(ranked, 1):
        if len(out_rows) >= top_n:
            break
        ck = f"{_normalize(row['name'])}::{category}"
        if ck in cache:
            v = cache[ck]
        else:
            result = verify_candidate(row, category, api_key)
            v = result.model_dump() if result else None
            cache[ck] = v
            VERIFY_CACHE_PATH.write_text(json.dumps(cache, indent=1, ensure_ascii=False))
            time.sleep(1)

        if not v or not v.get("identity_confirmed"):
            n_unconfirmed += 1
            log.info(f"  [{i}] {row['name']}: identity not confirmed, skipping")
            continue

        verdict = (v.get("recruitable") or "").strip().lower()
        if not (verdict.startswith("yes") or verdict.startswith("maybe")):
            n_unlikely += 1
            log.info(f"  [{i}] {row['name']}: {v.get('recruitable', 'no verdict')}, excluded")
            continue

        # Email policy: personal > academic > BLANK. Never a corporate
        # address — mailing someone at the employer we're recruiting them
        # away from is worse than no email. The hunt goes homepage (incl.
        # CV PDFs + de-obfuscation) -> GitHub commit emails -> their own
        # arXiv paper first pages.
        from email_hunt import hunt_personal_email, domain_kind
        email = v.get("personal_email", "")
        if email and domain_kind(email) in ("corporate", "junk"):
            email = ""
        if not email or domain_kind(email) != "personal":
            titles = [t.strip() for t in (row.get("key_papers") or "").split("|")]
            hunted = hunt_personal_email(row["name"], v.get("website", ""),
                                         titles, cache=hunt_cache)
            if hunted and (not email or domain_kind(hunted) == "personal"):
                email = hunted
        if not email and row.get("email") and domain_kind(row["email"]) == "academic":
            email = row["email"]

        out_rows.append({
            "#": 0,  # assigned after the verified-recruitability sort
            "Name": row["name"],
            "Career Stage": v.get("career_stage", ""),
            "Key Work": v.get("key_work", ""),
            "Personal Email": email,
            "Website": v.get("website", ""),
            "Recruitable?": v.get("recruitable", ""),
        })
        if i % 10 == 0:
            log.info(f"  [{category}] {i} examined, {len(out_rows)}/{top_n} slots filled")

    log.info(f"[{category}] examined {min(i, len(ranked)) if ranked else 0}: "
             f"{len(out_rows)} listed, {n_unlikely} unlikely (excluded), "
             f"{n_unconfirmed} unconfirmed (skipped)")

    # Order by the VERIFIED verdict (Yes before Maybe), then the walk order,
    # which already encodes the pipeline's hirability rank.
    def verdict_rank(r):
        return 0 if (r["Recruitable?"] or "").lower().startswith("yes") else 1
    out_rows.sort(key=verdict_rank)
    for n, r in enumerate(out_rows, 1):
        r["#"] = n

    out_path = out_dir / f"shortlist_{csv_path.stem}.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SHORTLIST_FIELDS)
        w.writeheader()
        w.writerows(out_rows)
    log.info(f"[{category}] shortlist written: {out_path} ({len(out_rows)} rows)")
    return out_rows


def write_shortlist_xlsx(tabs: dict, out_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in tabs.items():
        ws = wb.create_sheet(title=re.sub(r"[\[\]:*?/\\]", "_", title)[:31])
        ws.append(SHORTLIST_FIELDS)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for r in rows:
            ws.append([r.get(f, "") for f in SHORTLIST_FIELDS])
        ws.freeze_panes = "A2"
        widths = {"#": 4, "Name": 22, "Career Stage": 40, "Key Work": 55,
                  "Personal Email": 30, "Website": 28, "Recruitable?": 26}
        for i, f in enumerate(SHORTLIST_FIELDS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths[f]
    wb.save(out_path)
    log.info(f"Shortlist workbook: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Web-verify top candidates into recruiter-ready shortlists")
    parser.add_argument("--input-dir", default="data/expand_output")
    parser.add_argument("--top", type=int, default=40,
                        help="Confirmed Yes/Maybe rows per tab (not candidates examined)")
    parser.add_argument("--max-verify", type=int, default=0,
                        help="Cap on candidates examined per tab while filling slots (0 = 4x top)")
    parser.add_argument("--tabs", nargs="*", default=None,
                        help="Tab stems to process (default: every per-position CSV found)")
    args = parser.parse_args()

    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        log.error("OPENROUTER_API_KEY not set")
        return

    in_dir = Path(args.input_dir)
    csvs = sorted(p for p in in_dir.glob("*.csv")
                  if p.stem not in ("combined",) and not p.stem.startswith("shortlist_"))
    if args.tabs:
        csvs = [p for p in csvs if p.stem in set(args.tabs)]
    if not csvs:
        log.error(f"No per-position CSVs found in {in_dir}")
        return

    cache = _load_json(VERIFY_CACHE_PATH)
    hunt_cache = _load_json(HUNT_CACHE_PATH)
    tabs = {}
    for p in csvs:
        category = p.stem.replace("_", " ").title()
        tabs[category] = process_tab(p, category, args.top, api_key, cache, hunt_cache, in_dir,
                                     max_verify=args.max_verify)

    HUNT_CACHE_PATH.write_text(json.dumps(hunt_cache, indent=1, ensure_ascii=False))
    write_shortlist_xlsx(tabs, in_dir / "shortlist.xlsx")
    log.info("Done.")


if __name__ == "__main__":
    main()
