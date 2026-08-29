#!/usr/bin/env python3
"""
Seed-and-expand researcher discovery pipeline.

Starts from known papers/researchers, expands via citation graph (Semantic Scholar),
classifies by career stage and category fit (LLM), filters, enriches with emails.

Usage:
    python expand.py --seeds seeds.yaml --config config.yaml
    python expand.py --seeds seeds.yaml --config config.yaml --skip-emails
    python expand.py --seeds seeds.yaml --config config.yaml --dry-run
"""

import argparse
import csv
import json
import logging
import os
import re
import time
import xml.etree.ElementTree as ET
from collections import deque
from dataclasses import dataclass, field
from pathlib import Path

import requests
import yaml

from oa_client import OAClient
from scrape import _llm_call, _parse_llm_json, OPENROUTER_API, OPENROUTER_MODEL
from scrape_emails import scrape_emails as run_email_cascade, _normalize
from build_outreach_list import enrich_researcher as openalex_enrich

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent / "data"
EXPAND_CACHE_PATH = DATA_DIR / "expand_cache.json"

# ── Data structures ──────────────────────────────────────────────────────────

@dataclass
class Researcher:
    name: str
    s2_author_id: str = ""
    institution: str = ""
    country: str = ""
    homepage: str = ""
    h_index: int = 0
    cited_by_count: int = 0
    works_count: int = 0
    paper_count: int = 0
    key_papers: list = field(default_factory=list)
    career_stage: str = ""
    categories: list = field(default_factory=list)
    category_scores: dict = field(default_factory=dict)
    email: str = ""
    email_source: str = ""
    source_type: str = ""   # seed, coauthor, citation, reference, topic_search
    depth: int = 0
    found_via: str = ""     # provenance chain back to the seed, e.g. "GenSim2 <- cites <- cites"
    recruitable: str = "Yes"


def _cache_key(name: str) -> str:
    return _normalize(name)


def _extract_affiliation(author: dict) -> str:
    affiliations = author.get("affiliations") or []
    return affiliations[0] if affiliations else ""


def _load_json(path: Path) -> dict:
    if path.exists():
        try:
            return json.loads(path.read_text())
        except Exception:
            pass
    return {}


def _save_json(path: Path, data: dict):
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False))


# ── Phase 1: Resolve seeds ───────────────────────────────────────────────────

def _serialize_researchers(researchers: dict) -> list[dict]:
    """Convert researchers dict to JSON-serializable list."""
    out = []
    for key, r in researchers.items():
        out.append({
            "key": key, "name": r.name, "s2_author_id": r.s2_author_id,
            "institution": r.institution, "country": r.country, "homepage": r.homepage,
            "h_index": r.h_index, "cited_by_count": r.cited_by_count,
            "works_count": r.works_count, "paper_count": r.paper_count,
            "key_papers": r.key_papers[:10], "career_stage": r.career_stage,
            "categories": r.categories, "category_scores": r.category_scores,
            "email": r.email, "email_source": r.email_source,
            "source_type": r.source_type, "depth": r.depth,
            "found_via": r.found_via,
            "recruitable": r.recruitable,
        })
    return out


def _deserialize_researchers(data: list[dict]) -> dict:
    """Restore researchers dict from cached JSON."""
    researchers = {}
    for d in data:
        r = Researcher(
            name=d["name"], s2_author_id=d.get("s2_author_id", ""),
            institution=d.get("institution", ""), country=d.get("country", ""),
            homepage=d.get("homepage", ""),
            h_index=d.get("h_index", 0), cited_by_count=d.get("cited_by_count", 0),
            works_count=d.get("works_count", 0), paper_count=d.get("paper_count", 0),
            key_papers=d.get("key_papers", []), career_stage=d.get("career_stage", ""),
            categories=d.get("categories", []), category_scores=d.get("category_scores", {}),
            email=d.get("email", ""), email_source=d.get("email_source", ""),
            source_type=d.get("source_type", ""), depth=d.get("depth", 0),
            found_via=d.get("found_via", ""),
            recruitable=d.get("recruitable", "Yes"),
        )
        researchers[d["key"]] = r
    return researchers


def _seed_label(title: str) -> str:
    """Short human label for provenance chains: text before ':' or first 4 words."""
    if not title:
        return "?"
    head = title.split(":")[0].strip()
    if len(head) > 40:
        head = " ".join(head.split()[:4])
    return head[:40]


def resolve_seeds(seeds_cfg: dict, s2: OAClient) -> tuple[deque, dict]:
    """Resolve seed papers and researchers. Returns (queue, researchers)."""
    queue = deque()  # (paper_id, depth, source_type, via)
    researchers = {}
    seen_papers = set()

    # Seed papers (arxiv IDs)
    for arxiv_id in seeds_cfg.get("papers", []):
        paper = s2.paper_by_arxiv_id(arxiv_id)
        if paper and paper.get("paperId"):
            queue.append((paper["paperId"], 0, "seed", _seed_label(paper.get("title", arxiv_id))))
            seen_papers.add(paper["paperId"])
            log.info(f"  Seed paper: {paper.get('title', arxiv_id)[:80]}")
        else:
            log.warning(f"  Could not resolve arxiv:{arxiv_id}")

    # Seed papers (by title)
    for title in seeds_cfg.get("paper_titles", []):
        paper = s2.paper_by_title(title)
        if paper and paper.get("paperId"):
            queue.append((paper["paperId"], 0, "seed", _seed_label(paper.get("title", title))))
            seen_papers.add(paper["paperId"])
            log.info(f"  Seed paper (title): {paper.get('title', title)[:80]}")
        else:
            log.warning(f"  Could not resolve title: {title[:60]}")

    # Seed researchers
    for seed in seeds_cfg.get("researchers", []):
        name = seed["name"]
        institution = seed.get("institution", "")
        author = s2.search_author(name)
        if not author:
            log.warning(f"  Could not find researcher: {name}")
            continue

        key = _cache_key(name)
        researchers[key] = Researcher(
            name=name,
            s2_author_id=author.get("authorId", ""),
            institution=institution or _extract_affiliation(author),
            h_index=author.get("hIndex") or 0,
            cited_by_count=author.get("citationCount") or 0,
            paper_count=author.get("paperCount") or 0,
            source_type="seed",
            depth=0,
        )

        # Add their recent papers to queue
        if author.get("authorId"):
            papers = s2.get_author_papers(author["authorId"], limit=10)
            for p in papers:
                pid = p.get("paperId")
                if pid and pid not in seen_papers:
                    queue.append((pid, 0, "seed", f"{name} (seed researcher)"))
                    seen_papers.add(pid)
                    researchers[key].key_papers.append(p.get("title", ""))
        log.info(f"  Seed researcher: {name} ({institution}), {len(researchers[key].key_papers)} papers queued")

    return queue, researchers


# ── Phase 2: BFS expansion ───────────────────────────────────────────────────

def expand_graph(queue: deque, researchers: dict, s2: OAClient,
                 max_depth: int = 2, max_coauthors: int = 10,
                 max_citations: int = 50, max_references: int = 30,
                 prefer_recent: bool = True) -> set:
    """Level-synchronous BFS over the citation graph.

    Works are fetched in batches of 50 per OpenAlex call, which is what
    makes depth 2 viable: a 50K-paper leaf frontier costs ~1K calls
    (~10 min) instead of 50K (~7 h). References come embedded in the
    batched work objects, so only citation lookups are per-paper — and
    those exist only on non-leaf levels.

    Each queue entry carries a provenance string; researchers record it as
    found_via, e.g. "GenSim2 <- cites <- cited-by".

    Modifies researchers in-place. Returns seen_papers.
    """
    seen_papers = {entry[0] for entry in queue}
    level = list(queue)
    processed = 0

    while level:
        depth = level[0][1]
        ids = [entry[0] for entry in level]
        log.info(f"  Level {depth}: fetching {len(ids)} papers "
                 f"({(len(ids) + 49) // 50} batch calls)")

        works = {}
        for i in range(0, len(ids), 50):
            works.update(s2.get_works_batch(ids[i:i + 50]))

        next_level = []
        for paper_id, d, source_type, via in level:
            processed += 1
            if processed % 500 == 0:
                log.info(f"  Processed {processed} papers, {len(researchers)} researchers")

            paper_info = works.get(paper_id) or {}
            authors = paper_info.get("authors") or []
            paper_title = paper_info.get("title", "")

            # Authors from BOTH ends: ML papers put PIs last; taking only
            # authors[:N] silently dropped exactly the senior authors whose
            # networks we want.
            if len(authors) > max_coauthors:
                half = max_coauthors // 2
                selected = authors[:max_coauthors - half] + authors[-half:]
            else:
                selected = authors

            for author in selected:
                name = author.get("name", "")
                if not name or len(name) < 3:
                    continue
                key = _cache_key(name)
                if key not in researchers:
                    researchers[key] = Researcher(
                        name=name,
                        s2_author_id=author.get("authorId", ""),
                        institution=_extract_affiliation(author),
                        h_index=author.get("hIndex") or 0,
                        cited_by_count=author.get("citationCount") or 0,
                        paper_count=author.get("paperCount") or 0,
                        key_papers=[paper_title] if paper_title else [],
                        source_type=source_type if source_type != "seed" else "coauthor",
                        depth=d,
                        found_via=via,
                    )
                else:
                    if paper_title and paper_title not in researchers[key].key_papers:
                        researchers[key].key_papers.append(paper_title)
                    if d < researchers[key].depth:
                        researchers[key].depth = d
                        researchers[key].found_via = via

            if d < max_depth:
                # Citations still need one call per paper (can't batch
                # different cite-targets into one filter).
                citations = s2.get_citations(paper_id, limit=max_citations)
                if prefer_recent:
                    citations.sort(key=lambda c: c.get("year") or 0, reverse=True)
                for cit in citations[:max_citations]:
                    cit_id = cit.get("paperId")
                    if cit_id and cit_id not in seen_papers:
                        seen_papers.add(cit_id)
                        next_level.append((cit_id, d + 1, "citation", f"{via} <- cites"))

                # References are embedded in the batched work — no extra call.
                refs = (paper_info.get("referenced_works") or [])[:max_references]
                for ref_id in refs:
                    if ref_id and ref_id not in seen_papers:
                        seen_papers.add(ref_id)
                        next_level.append((ref_id, d + 1, "reference", f"{via} <- cited-by"))

        level = next_level

    log.info(f"  BFS done: {processed} papers processed, {len(researchers)} researchers found")
    return seen_papers


# ── Phase 2b: Topic search expansion ─────────────────────────────────────────

def expand_via_topics(config: dict, s2: OAClient, researchers: dict, seen_papers: set):
    """Keyword search on S2 to find additional researchers."""
    categories = config.get("categories", [])
    topic_limit = config.get("expansion", {}).get("topic_search_limit", 50)

    for cat in categories:
        for kw in cat.get("keywords", []):
            papers = s2.search_papers(kw, year_range="2024-2026", limit=topic_limit)
            added = 0
            for p in papers:
                pid = p.get("paperId")
                if not pid or pid in seen_papers:
                    continue
                seen_papers.add(pid)
                for author in (p.get("authors") or []):
                    name = author.get("name", "")
                    if not name or len(name) < 3:
                        continue
                    key = _cache_key(name)
                    if key not in researchers:
                        researchers[key] = Researcher(
                            name=name,
                            s2_author_id=author.get("authorId", ""),
                            key_papers=[p.get("title", "")],
                            source_type="topic_search",
                            depth=1,
                        )
                        added += 1
                    elif p.get("title") and p["title"] not in researchers[key].key_papers:
                        researchers[key].key_papers.append(p["title"])
            log.info(f"  Topic '{kw}': {added} new researchers")


# ── Phase 3: LLM career stage classification ─────────────────────────────────

CAREER_STAGE_PROMPT = """Classify each ML researcher by career stage. You have: name, institution, h-index, paper count, and recent paper titles.

Categories (pick ONE per person):
- graduating_phd: PhD student in 4th+ year, likely graduating 2025-2026
- recent_grad: Graduated PhD 2024-2025, now in first industry role or postdoc
- junior_industry: 2-6 years post-PhD at a company (not a professor)
- mid_industry: 6-10 years experience, IC/senior IC role at a company
- senior: 10+ years, staff/principal level
- professor: Faculty at a university
- early_phd: 1st-3rd year PhD student
- founder: CEO/CTO/co-founder of a company
- unknown: Cannot determine

Respond with ONLY a JSON array: [{"id": 1, "stage": "graduating_phd"}, ...]. No explanation.

Researchers:
"""


def classify_career_stages(researchers: dict, api_key: str, cache: dict):
    """Batch-classify career stages via LLM. Modifies researchers in-place.

    Cache key is versioned: v1 labels were produced BEFORE enrichment, when
    the prompt said "Institution: unknown, h-index: 0" for everyone — 71%
    came back "unknown". v2 labels are classified on enriched profiles.
    """
    cached_stages = cache.get("career_stages_v2", {})
    to_classify = []
    for key, r in researchers.items():
        if key in cached_stages:
            r.career_stage = cached_stages[key]
        elif not r.career_stage:
            to_classify.append((key, r))

    if not to_classify:
        return
    log.info(f"  Classifying {len(to_classify)} researchers ({len(cached_stages)} cached)")

    batch_size = 20
    for i in range(0, len(to_classify), batch_size):
        batch = to_classify[i:i + batch_size]
        lines = []
        for j, (key, r) in enumerate(batch):
            papers_str = " | ".join(r.key_papers[:5])
            lines.append(f"{j+1}. Name: {r.name}, Institution: {r.institution or 'unknown'}, "
                         f"h-index: {r.h_index}, Papers: {r.paper_count}, "
                         f"Recent: {papers_str[:200]}")

        content = _llm_call(api_key, CAREER_STAGE_PROMPT + "\n".join(lines))
        if content:
            try:
                results = _parse_llm_json(content)
                for item in results:
                    idx = item.get("id", 0) - 1
                    if 0 <= idx < len(batch):
                        key, r = batch[idx]
                        r.career_stage = item.get("stage", "unknown")
                        cached_stages[key] = r.career_stage
            except (json.JSONDecodeError, KeyError, TypeError) as e:
                log.warning(f"  Career stage parse error: {e}")

        batch_num = i // batch_size + 1
        total_batches = (len(to_classify) + batch_size - 1) // batch_size
        log.info(f"  Career stage batch {batch_num}/{total_batches}")
        time.sleep(0.5)

    cache["career_stages_v2"] = cached_stages


# ── Phase 4: LLM category fit classification ─────────────────────────────────

CATEGORY_FIT_PROMPT = """Rate how well each researcher fits each category based on their recent paper titles.

Categories:
{categories_block}

Score each researcher 0.0-1.0 per category (0 = no fit, 1 = perfect fit).

Respond with ONLY a JSON array: [{{"id": 1, "scores": {{"Category A": 0.8, "Category B": 0.1}}}}]. No explanation.

Researchers:
"""


def classify_categories(researchers: dict, categories: list, api_key: str, cache: dict):
    """Batch-classify category fit via LLM. Modifies researchers in-place."""
    cached_fit = cache.get("category_fit", {})
    cat_names = [c["name"] for c in categories]
    categories_block = "\n".join(f"- {c['name']}: {', '.join(c.get('keywords', []))}" for c in categories)

    to_classify = []
    for key, r in researchers.items():
        # A cache hit is only valid if it covers every CURRENT category;
        # otherwise researchers cached before a category was added could
        # never join it (hit with the "Benchmarks" role, 2026-08-29).
        if key in cached_fit and all(n in cached_fit[key] for n in cat_names):
            r.category_scores = cached_fit[key]
            r.categories = [name for name, score in cached_fit[key].items() if score >= 0.5]
        else:
            to_classify.append((key, r))

    if not to_classify:
        return
    log.info(f"  Classifying {len(to_classify)} researchers into {len(categories)} categories ({len(cached_fit)} cached)")

    batch_size = 15
    for i in range(0, len(to_classify), batch_size):
        batch = to_classify[i:i + batch_size]
        lines = []
        for j, (key, r) in enumerate(batch):
            papers_str = " | ".join(r.key_papers[:8])
            lines.append(f"{j+1}. {r.name}: {papers_str[:250]}")

        prompt = CATEGORY_FIT_PROMPT.format(categories_block=categories_block) + "\n".join(lines)
        content = _llm_call(api_key, prompt)
        if content:
            try:
                results = _parse_llm_json(content)
                for item in results:
                    idx = item.get("id", 0) - 1
                    if 0 <= idx < len(batch):
                        key, r = batch[idx]
                        scores = item.get("scores", {})
                        r.category_scores = scores
                        r.categories = [name for name, score in scores.items() if score >= 0.5]
                        cached_fit[key] = scores
            except (json.JSONDecodeError, KeyError, TypeError) as e:
                log.warning(f"  Category fit parse error: {e}")

        batch_num = i // batch_size + 1
        total_batches = (len(to_classify) + batch_size - 1) // batch_size
        log.info(f"  Category fit batch {batch_num}/{total_batches}")
        time.sleep(0.5)

    cache["category_fit"] = cached_fit


# ── Phase 5: Filtering ────────────────────────────────────────────────────────

def apply_filters(researchers: dict, config: dict, exclude_names: set) -> dict:
    """Apply configured filters. Returns filtered dict.

    Runs AFTER ID-based enrichment, so h_index / institution / country are
    real. Policy: hard-drop only what we can never hire (wrong geography,
    excluded orgs, no category fit, too-early PhDs); seniors and h>=cap
    people are KEPT but flagged recruitable="Stretch" — a founder or PI who
    is interested is very recruitable, and dropping them silently is how we
    lost the LAB-Bench PI from the sheet.
    """
    filters = config.get("filters", {})
    max_h = filters.get("max_h_index", 40)
    min_papers = filters.get("min_papers", 1)
    allowed_stages = set(filters.get("career_stages", [
        "graduating_phd", "recent_grad", "junior_industry", "mid_industry"
    ]))
    exclude_institutions = {i.lower() for i in filters.get("exclude_institutions", [])}
    user_exclude_names = {n.lower() for n in filters.get("exclude_names", [])}
    allowed_countries = {c.upper() for c in filters.get("allowed_countries", [])}
    drop_stages = set(filters.get("drop_stages", ["early_phd"]))

    filtered = {}
    removed = {"papers": 0, "stage": 0, "institution": 0, "country": 0,
               "name": 0, "category": 0, "dedup": 0}
    flagged = {"stretch": 0, "unlikely": 0}

    hard_recruit = {"openai", "anthropic", "google deepmind", "deepmind"}

    for key, r in researchers.items():
        if key in exclude_names or r.name.lower() in user_exclude_names:
            removed["dedup"] += 1
            continue
        if r.paper_count and r.paper_count < min_papers:
            removed["papers"] += 1
            continue
        if r.career_stage in drop_stages:
            removed["stage"] += 1
            continue
        if any(exc in (r.institution or "").lower() for exc in exclude_institutions):
            removed["institution"] += 1
            continue
        # Geography: drop only when the country is KNOWN and disallowed.
        # Unknown country (no institution data) stays in — dropping on
        # missing data would cut ~30% of the sheet for no signal.
        if allowed_countries and r.country and r.country.upper() not in allowed_countries:
            removed["country"] += 1
            continue
        if not r.categories:
            removed["category"] += 1
            continue

        # Flag, don't drop
        if any(hr in (r.institution or "").lower() for hr in hard_recruit):
            r.recruitable = "Unlikely"
            flagged["unlikely"] += 1
        elif (r.career_stage in ("senior", "professor", "founder")
              or (r.h_index and r.h_index >= max_h)):
            r.recruitable = "Stretch"
            flagged["stretch"] += 1
        elif r.career_stage in allowed_stages or r.career_stage == "unknown" or not r.career_stage:
            r.recruitable = "Yes"

        filtered[key] = r

    log.info(f"  Filtered: {len(researchers)} -> {len(filtered)} "
             f"(removed: {removed}, flagged: {flagged})")
    return filtered


# ── Phase 6: Enrichment ───────────────────────────────────────────────────────

def enrich_profiles_by_id(researchers: dict):
    """Batch OpenAlex profile fetch by author ID for ALL researchers.

    The ID was captured from the paper's authorships during the graph walk,
    so it names the exact person — immune to the display-name collisions
    that a search by name suffers (e.g. the wrong high-h "Chao Wang").
    Runs BEFORE classification: at ~50 profiles per call the whole graph
    enriches in about a minute, and the career-stage LLM then sees real
    h-index / institution / country instead of zeros (which previously made
    it label 71% of people "unknown").
    """
    from oa_client import OAClient
    oa = OAClient()

    id_cache = _load_json(DATA_DIR / "enrich_cache_by_id.json")
    with_id = [(k, r) for k, r in researchers.items()
               if r.s2_author_id and r.s2_author_id.startswith("A")]
    to_fetch = [r.s2_author_id for k, r in with_id if r.s2_author_id not in id_cache]
    log.info(f"  ID-based enrichment: {len(with_id)} researchers with OpenAlex IDs "
             f"({len(id_cache)} cached, {len(to_fetch)} to fetch)")

    for i in range(0, len(to_fetch), 50):
        batch = to_fetch[i:i + 50]
        profiles = oa.get_authors_batch(batch)
        for aid in batch:
            id_cache[aid] = profiles.get(aid, {})
        if (i // 50) % 10 == 0:
            _save_json(DATA_DIR / "enrich_cache_by_id.json", id_cache)
            log.info(f"  Fetched {min(i + 50, len(to_fetch))}/{len(to_fetch)} profiles...")
    _save_json(DATA_DIR / "enrich_cache_by_id.json", id_cache)

    id_enriched = 0
    for key, r in with_id:
        p = id_cache.get(r.s2_author_id) or {}
        if not p:
            continue
        # Sanity: the ID must belong to this person. Graph-walk IDs always do,
        # but seed researchers got theirs from a name search which can top-hit
        # the wrong profile. Require one shared name token; otherwise leave the
        # row to the name-based fallback.
        r_tokens = {t.lower() for t in r.name.replace(",", " ").split() if len(t) > 2}
        p_tokens = {t.lower() for t in (p.get("name") or "").replace(",", " ").split() if len(t) > 2}
        if r_tokens and p_tokens and not (r_tokens & p_tokens):
            log.warning(f"  ID/name mismatch, skipping ID-enrich: {r.name!r} vs {p.get('name')!r}")
            continue
        # Authoritative: overwrite, don't merge
        r.h_index = p.get("hIndex") or 0
        r.cited_by_count = p.get("citationCount") or 0
        r.works_count = p.get("paperCount") or 0
        if not r.paper_count:
            r.paper_count = p.get("paperCount") or 0
        affs = p.get("affiliations") or []
        if affs:
            r.institution = affs[0]
        if p.get("country"):
            r.country = p["country"]
        id_enriched += 1
    log.info(f"  ID-based enrichment done: {id_enriched}/{len(with_id)}")


def enrich(researchers: dict, skip_emails: bool = False):
    """Post-filter enrichment: name-search fallback for rows the ID pass
    missed, then the email cascade. Runs only on filter survivors."""
    enrich_cache = _load_json(DATA_DIR / "enrich_cache.json")
    name_enriched = 0
    for key, r in researchers.items():
        if r.h_index == 0 or not r.institution:
            result = openalex_enrich(r.name, r.institution, enrich_cache)
            if result:
                r.h_index = result.get("h_index", r.h_index) or r.h_index
                r.cited_by_count = result.get("cited_by_count", r.cited_by_count) or r.cited_by_count
                r.works_count = result.get("works_count", r.works_count) or r.works_count
                if not r.paper_count:
                    r.paper_count = result.get("works_count", 0) or r.paper_count
                if not r.institution:
                    r.institution = result.get("institution", "") or r.institution
                if not r.homepage and result.get("homepage"):
                    r.homepage = result["homepage"]
                name_enriched += 1
    _save_json(DATA_DIR / "enrich_cache.json", enrich_cache)
    log.info(f"  Name-based fallback enriched: {name_enriched}")

    if skip_emails:
        return

    # Email cascade — convert to row format
    rows = [{"name": r.name, "institution": r.institution or "",
             "paper_count": str(r.paper_count), "priority_score": "0"}
            for r in researchers.values()]

    enriched_rows = run_email_cascade(rows)

    # Map back
    for row in enriched_rows:
        key = _cache_key(row["name"])
        if key in researchers:
            researchers[key].email = row.get("email", "")
            researchers[key].email_source = row.get("email_source", "")
            researchers[key].homepage = row.get("homepage", "")

    with_email = sum(1 for r in researchers.values() if r.email)
    log.info(f"  Emails found: {with_email}/{len(researchers)}")


# ── Phase 7: Output ───────────────────────────────────────────────────────────

OUTPUT_FIELDS = [
    "name", "career_stage", "institution", "country", "categories", "key_papers",
    "email", "email_source", "homepage", "h_index", "cited_by_count",
    "paper_count", "source_type", "depth", "found_via", "recruitable",
]

_RECRUIT_ORDER = {"Yes": 0, "Stretch": 1, "Unlikely": 2}


def write_output(researchers: dict, categories: list, output_dir: Path) -> tuple[Path, dict, list[dict]]:
    """Write per-category CSVs, combined.csv, and a single xlsx with a tab per
    position. Returns (xlsx_path, per_cat_rows, all_rows) so callers (Slack
    post) can reuse the row data without re-building it."""
    output_dir.mkdir(parents=True, exist_ok=True)

    all_rows = []
    # Recruitable Yes first, then Stretch, then Unlikely; h-index desc within
    for key, r in sorted(researchers.items(),
                         key=lambda x: (_RECRUIT_ORDER.get(x[1].recruitable, 0), -x[1].h_index)):
        row = {
            "name": r.name,
            "career_stage": r.career_stage,
            "institution": r.institution,
            "country": r.country or "?",
            "categories": "; ".join(r.categories),
            "key_papers": " | ".join(r.key_papers[:5]),
            "email": r.email,
            "email_source": r.email_source,
            "homepage": r.homepage,
            "h_index": r.h_index,
            "cited_by_count": r.cited_by_count,
            "paper_count": r.paper_count,
            "source_type": r.source_type,
            "depth": r.depth,
            "found_via": r.found_via,
            "recruitable": r.recruitable,
        }
        all_rows.append(row)

    # CSVs
    _write_csv(output_dir / "combined.csv", all_rows)
    log.info(f"  combined.csv: {len(all_rows)} researchers")

    per_cat_rows = {}
    for cat in categories:
        cat_name = cat["name"]
        # Exact membership: substring matching put every "Agentic/STEM
        # Benchmarks" member into the "Benchmarks" tab.
        cat_rows = [r for r in all_rows if cat_name in r["categories"].split("; ")]
        per_cat_rows[cat_name] = cat_rows
        safe_name = re.sub(r"[^a-z0-9_]", "_", cat_name.lower())
        _write_csv(output_dir / f"{safe_name}.csv", cat_rows)
        log.info(f"  {safe_name}.csv: {len(cat_rows)} researchers")

    # XLSX: one tab per position, plus a Combined tab only when >1 position
    xlsx_path = output_dir / "researchers.xlsx"
    _write_xlsx(xlsx_path, per_cat_rows, all_rows)
    extra = " + Combined" if len(per_cat_rows) > 1 else ""
    log.info(f"  {xlsx_path.name}: {len(per_cat_rows)} position tab(s){extra}")
    return xlsx_path, per_cat_rows, all_rows


def _write_csv(path: Path, rows: list[dict]):
    if not rows:
        return
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, per_cat_rows: dict, all_rows: list[dict]):
    """Write one .xlsx: a Combined tab, then one tab per position (category)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    def _add_sheet(title: str, rows: list[dict]):
        # Excel sheet-name limit is 31 chars and disallows []:*?/\
        safe = re.sub(r"[\[\]:*?/\\]", "_", title)[:31] or "Sheet"
        ws = wb.create_sheet(title=safe)
        ws.append(OUTPUT_FIELDS)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
        for r in rows:
            ws.append([r.get(f, "") for f in OUTPUT_FIELDS])
        ws.freeze_panes = "A2"
        # Auto-ish column widths
        for i, f in enumerate(OUTPUT_FIELDS, start=1):
            width = max(12, min(60, max((len(str(r.get(f, ""))) for r in rows), default=len(f)) + 2))
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # With a single category the Combined tab is an identical copy of the one
    # position tab — emit just the position tab in that case.
    if len(per_cat_rows) > 1:
        _add_sheet("Combined", all_rows)
    for cat_name, rows in per_cat_rows.items():
        _add_sheet(cat_name, rows)

    wb.save(path)


# ── Dedup helpers ─────────────────────────────────────────────────────────────

def load_existing_names(csv_path: str) -> set:
    """Load normalized names from an existing CSV for deduplication."""
    names = set()
    path = Path(csv_path)
    if not path.exists():
        return names
    with open(path) as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("name", "")
            if name:
                names.add(_cache_key(name))
    return names


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed-and-expand researcher discovery")
    parser.add_argument("--seeds", required=True, help="Path to seeds YAML")
    parser.add_argument("--config", required=True, help="Path to config YAML")
    parser.add_argument("--max-depth", type=int, default=None,
                        help="Override max expansion depth (0=seeds only, 1=one hop, 2=two hops)")
    parser.add_argument("--skip-emails", action="store_true", help="Skip email discovery")
    parser.add_argument("--skip-enrichment", action="store_true",
                        help="Skip OpenAlex profile enrichment (h-index, institution). "
                             "Useful when OpenAlex is 429-ing hard; xlsx will have h=0 for uncached authors.")
    parser.add_argument("--skip-classify", action="store_true", help="Skip LLM classification")
    parser.add_argument("--skip-topics", action="store_true", help="Skip topic keyword search")
    parser.add_argument("--dry-run", action="store_true", help="Expand only, skip classify/enrich/email")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from cached state (skip seed resolution + BFS, go straight to classify/filter/enrich)")
    parser.add_argument("--output-dir", default="data/expand_output", help="Output directory")
    parser.add_argument("--dedup-csv", default="", help="CSV to dedup against")
    parser.add_argument("--slack-channel", default="",
                        help="Slack channel name or ID (e.g. 'hiring'). If set and "
                             "SLACK_BOT_TOKEN is in env, uploads the xlsx and posts a summary.")
    parser.add_argument("--cache", default=str(EXPAND_CACHE_PATH),
                        help="Path for the expansion/LLM-label cache. Use a separate file "
                             "for exploratory runs so the main pipeline's cache is untouched.")
    parser.add_argument("--min-key-papers", type=int, default=0,
                        help="Before LLM classification, drop researchers found at depth>=2 "
                             "with fewer than N in-graph papers. Depth-2 frontiers are huge "
                             "and dominated by one-off co-authors; N=2 keeps the LLM bill sane.")
    args = parser.parse_args()

    with open(args.seeds) as f:
        seeds_cfg = yaml.safe_load(f)["seeds"]
    with open(args.config) as f:
        config = yaml.safe_load(f)

    openrouter_key = os.environ.get("OPENROUTER_API_KEY", "")
    s2 = OAClient()
    cache_path = Path(args.cache)
    expand_cache = _load_json(cache_path)

    # Depth override
    exp = config.get("expansion", {})
    max_depth = args.max_depth if args.max_depth is not None else exp.get("max_depth", 2)

    # Load existing names for dedup
    exclude_names = set()
    if args.dedup_csv:
        exclude_names = load_existing_names(args.dedup_csv)
    log.info(f"Dedup against {len(exclude_names)} existing names")

    # ── Resume from cache or expand fresh ──
    if args.resume and "researchers" in expand_cache:
        researchers = _deserialize_researchers(expand_cache["researchers"])
        seen_papers = set(expand_cache.get("seen_papers", []))
        log.info(f"=== Resumed: {len(researchers)} researchers, {len(seen_papers)} seen papers ===")
    else:
        # Phase 1: Resolve seeds
        log.info("=== Phase 1: Resolve Seeds ===")
        queue, researchers = resolve_seeds(seeds_cfg, s2)
        log.info(f"  Seeds resolved: {len(queue)} papers queued, {len(researchers)} seed researchers")

        # Phase 2: BFS expansion
        log.info(f"=== Phase 2: Graph Expansion (depth={max_depth}) ===")
        seen_papers = expand_graph(
            queue, researchers, s2,
            max_depth=max_depth,
            max_coauthors=exp.get("max_coauthors_per_paper", 10),
            max_citations=exp.get("max_citations_per_paper", 50),
            max_references=exp.get("max_references_per_paper", 30),
            prefer_recent=exp.get("prefer_recent", True),
        )

        # Phase 2b: Topic search
        if not args.skip_topics:
            log.info("=== Phase 2b: Topic Search ===")
            expand_via_topics(config, s2, researchers, seen_papers)

        log.info(f"  Total researchers after expansion: {len(researchers)}")

        # Save expansion state for resume
        expand_cache["researchers"] = _serialize_researchers(researchers)
        expand_cache["seen_papers"] = list(seen_papers)
        expand_cache["last_depth"] = max_depth
        _save_json(cache_path, expand_cache)
        log.info(f"  Expansion state saved ({len(researchers)} researchers, depth={max_depth})")

    if args.dry_run:
        log.info("Dry run — writing raw expansion output")
        write_output(researchers, config.get("categories", []), Path(args.output_dir))
        return

    # Deep-frontier gate: at depth>=2 the frontier is dominated by one-off
    # co-authors of tangentially-related papers. Requiring >=N in-graph
    # papers before spending LLM tokens on someone keeps classification
    # tractable without losing anyone with a real footprint in the topic.
    if args.min_key_papers > 1:
        before = len(researchers)
        researchers = {k: r for k, r in researchers.items()
                       if r.depth < 2 or len(r.key_papers) >= args.min_key_papers}
        log.info(f"  Deep-frontier gate (depth>=2 needs {args.min_key_papers}+ in-graph papers): "
                 f"{before} -> {len(researchers)}")

    # Phase 3: Profile enrichment by author ID — BEFORE classification, so
    # the career-stage LLM sees real h-index / institution / country instead
    # of zeros. ~1 minute for the whole graph at 50 profiles per call.
    if not args.skip_enrichment:
        log.info("=== Phase 3: Profile Enrichment (by author ID) ===")
        enrich_profiles_by_id(researchers)
        # Persist enriched fields into the resume cache
        expand_cache["researchers"] = _serialize_researchers(researchers)
        _save_json(cache_path, expand_cache)

    # Phase 4: Career stage classification
    if not args.skip_classify and openrouter_key:
        log.info("=== Phase 4: Career Stage Classification ===")
        classify_career_stages(researchers, openrouter_key, expand_cache)
        _save_json(cache_path, expand_cache)

    # Phase 5: Category classification
    if not args.skip_classify and openrouter_key:
        log.info("=== Phase 5: Category Classification ===")
        classify_categories(researchers, config.get("categories", []), openrouter_key, expand_cache)
        _save_json(cache_path, expand_cache)

    # Phase 6: Filter (runs on enriched values; flags Stretch/Unlikely
    # instead of dropping seniors — see apply_filters)
    log.info("=== Phase 6: Filtering ===")
    researchers = apply_filters(researchers, config, exclude_names)

    # Phase 6b: Fallback enrichment + emails, survivors only
    if args.skip_enrichment:
        log.info("=== Phase 6b: Fallback enrichment SKIPPED (--skip-enrichment) ===")
    else:
        log.info("=== Phase 6b: Fallback Enrichment + Emails ===")
        enrich(researchers, skip_emails=args.skip_emails)

    # Phase 7: Output
    log.info("=== Phase 7: Output ===")
    xlsx_path, per_cat_rows, all_rows = write_output(
        researchers, config.get("categories", []), Path(args.output_dir)
    )
    _save_json(cache_path, expand_cache)

    # Phase 8: Slack (optional)
    slack_token = os.environ.get("SLACK_BOT_TOKEN", "")
    if args.slack_channel and slack_token:
        log.info(f"=== Phase 8: Slack post to #{args.slack_channel.lstrip('#')} ===")
        from slack_post import upload_file, build_summary
        summary = build_summary(per_cat_rows, all_rows)
        result = upload_file(str(xlsx_path), args.slack_channel, slack_token,
                             initial_comment=summary)
        if result:
            log.info(f"  Uploaded {xlsx_path.name} to Slack (file id: {result.get('id')})")
        else:
            log.warning("  Slack upload failed — see errors above")
    elif args.slack_channel and not slack_token:
        log.warning("--slack-channel set but SLACK_BOT_TOKEN missing; skipping")

    log.info("Done.")


if __name__ == "__main__":
    main()
